import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import database as db

def apply_header_style(cell, text, bg_color="1E3A8A", text_color="FFFFFF"):
    cell.value = text
    cell.font = Font(name="Arial", size=11, bold=True, color=text_color)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def apply_row_style(cell, value, align="left", is_money=False):
    cell.value = value
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if is_money and isinstance(value, (int, float)):
        cell.number_format = '#,##0 "so''m"'

async def generate_shop_excel(shop_id: int, ledger_type: str = 'receivable') -> io.BytesIO:
    """Bitta do'kon uchun alohida to'liq Excel hisoboti (So'm va Dollar)"""
    shop = await db.get_shop_by_id(shop_id)
    customers = await db.get_customers_by_shop(shop_id, ledger_type=ledger_type)
    
    wb = openpyxl.Workbook()
    is_payable = (ledger_type == 'payable')
    
    # 1-Varaq: Shaxslar va Qarzlar
    ws_cust = wb.active
    ws_cust.title = "Haqdorlar va Qarzlar" if is_payable else "Qarzdorlar va Qarzlar"
    
    person_header = "Haqdor (Qarz beruvchi)" if is_payable else "Qarzdor / Mijoz Ismi"
    headers = ["№", person_header, "Telefon Raqami", "So'm Qarzi", "Dollar Qarzi ($)", "Telegram Ulangan"]
    for col_idx, h in enumerate(headers, 1):
        apply_header_style(ws_cust.cell(row=1, column=col_idx), h, bg_color="2563EB")
        
    for row_idx, c in enumerate(customers, 2):
        apply_row_style(ws_cust.cell(row=row_idx, column=1), row_idx - 1, align="center")
        apply_row_style(ws_cust.cell(row=row_idx, column=2), c['full_name'])
        apply_row_style(ws_cust.cell(row=row_idx, column=3), c['phone'] or "-", align="center")
        apply_row_style(ws_cust.cell(row=row_idx, column=4), c.get('balance', 0.0) or 0.0, align="right", is_money=True)
        
        usd_cell = ws_cust.cell(row=row_idx, column=5)
        usd_val = c.get('balance_usd', 0.0) or 0.0
        usd_cell.value = usd_val
        usd_cell.font = Font(name="Arial", size=10)
        usd_cell.alignment = Alignment(horizontal="right", vertical="center")
        usd_cell.number_format = '#,##0.00 "$" '
        
        apply_row_style(ws_cust.cell(row=row_idx, column=6), "Ha" if c.get('telegram_id') else "Yo'q", align="center")
        
    # Kengliklarni moslash
    for col in ws_cust.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_cust.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)
        
    # 2-Varaq: Savdolar tarixi
    ws_tx = wb.create_sheet(title="Amallar Tarixi")
    tx_headers = ["№", "Sana va Vaqt", "Mijoz Ismi", "Amal Turi", "Summa", "Valyuta", "Izoh"]
    for col_idx, h in enumerate(tx_headers, 1):
        apply_header_style(ws_tx.cell(row=1, column=col_idx), h, bg_color="0D9488")
        
    row_num = 2
    for c in customers:
        txs = await db.get_customer_transactions(c['id'], limit=100)
        for t in txs:
            t_type = "🔴 Qarz berildi" if t['type'] == 'debt' else "🟢 To'lov qilindi"
            curr = t.get('currency', 'UZS') or 'UZS'
            apply_row_style(ws_tx.cell(row=row_num, column=1), row_num - 1, align="center")
            apply_row_style(ws_tx.cell(row=row_num, column=2), str(t['created_at'])[:19], align="center")
            apply_row_style(ws_tx.cell(row=row_num, column=3), c['full_name'])
            apply_row_style(ws_tx.cell(row=row_num, column=4), t_type, align="center")
            
            s_cell = ws_tx.cell(row=row_num, column=5)
            s_cell.value = t['amount']
            s_cell.font = Font(name="Arial", size=10)
            s_cell.alignment = Alignment(horizontal="right", vertical="center")
            s_cell.number_format = '#,##0.00' if curr == 'USD' else '#,##0'
            
            apply_row_style(ws_tx.cell(row=row_num, column=6), curr, align="center")
            apply_row_style(ws_tx.cell(row=row_num, column=7), t['description'] or "-")
            row_num += 1
            
    for col in ws_tx.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_tx.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

async def generate_full_platform_excel() -> io.BytesIO:
    """Super Admin uchun butun platformaning to'liq zaxira Excel fayli (Do'konlar, Qarzdor/Haqdorlar, Amallar tarixi)"""
    shops = await db.get_detailed_shops_analysis()
    
    wb = openpyxl.Workbook()
    
    # 1-Varaq: Barcha Do'konlar va Daftarlar
    ws_shops = wb.active
    ws_shops.title = "Do'konlar Ro'yxati"
    
    headers = ["ID", "Do'kon Nomi", "Admin Telegram ID", "Telefon", "Mijozlar Soni", "Holat", "Obuna Qolgan Kun"]
    for col_idx, h in enumerate(headers, 1):
        apply_header_style(ws_shops.cell(row=1, column=col_idx), h, bg_color="1E3A8A")
        
    for row_idx, s in enumerate(shops, 2):
        status = "Faol" if s['is_active'] else "Bloklangan"
        apply_row_style(ws_shops.cell(row=row_idx, column=1), s['id'], align="center")
        apply_row_style(ws_shops.cell(row=row_idx, column=2), s['name'])
        apply_row_style(ws_shops.cell(row=row_idx, column=3), s['admin_id'], align="center")
        apply_row_style(ws_shops.cell(row=row_idx, column=4), s['phone'] or "-", align="center")
        apply_row_style(ws_shops.cell(row=row_idx, column=5), s.get('customers_count', 0), align="center")
        apply_row_style(ws_shops.cell(row=row_idx, column=6), status, align="center")
        apply_row_style(ws_shops.cell(row=row_idx, column=7), s.get('days_left', 30), align="center")
        
    for col in ws_shops.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_shops.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)
        
    # 2-Varaq: Barcha Qarzdorlar va Haqdorlar
    ws_all_cust = wb.create_sheet(title="Qarzdorlar va Haqdorlar")
    cust_headers = ["ID", "Do'kon Nomi", "Shaxs Ismi", "Turi", "Telefon", "So'm Qarzi", "Dollar Qarzi ($)", "Muddat Sanasi", "Telegram ID"]
    for col_idx, h in enumerate(cust_headers, 1):
        apply_header_style(ws_all_cust.cell(row=1, column=col_idx), h, bg_color="4338CA")
        
    c_row = 2
    for s in shops:
        c_list = await db.get_customers_by_shop(s['id'])
        for c in c_list:
            l_type = "🔴 Haqdor (Berishim kerak)" if c.get('ledger_type') == 'payable' else "🟢 Qarzdor (Olishim kerak)"
            due_str = str(c.get('due_date'))[:10] if c.get('due_date') else "-"
            u_bal = c.get('balance', 0.0) or 0.0
            d_bal = c.get('balance_usd', 0.0) or 0.0
            
            apply_row_style(ws_all_cust.cell(row=c_row, column=1), c['id'], align="center")
            apply_row_style(ws_all_cust.cell(row=c_row, column=2), s['name'])
            apply_row_style(ws_all_cust.cell(row=c_row, column=3), c['full_name'])
            apply_row_style(ws_all_cust.cell(row=c_row, column=4), l_type, align="center")
            apply_row_style(ws_all_cust.cell(row=c_row, column=5), c['phone'] or "-", align="center")
            apply_row_style(ws_all_cust.cell(row=c_row, column=6), u_bal, align="right", is_money=True)
            
            d_cell = ws_all_cust.cell(row=c_row, column=7)
            d_cell.value = d_bal
            d_cell.font = Font(name="Arial", size=10)
            d_cell.alignment = Alignment(horizontal="right", vertical="center")
            d_cell.number_format = '#,##0.00'
            
            apply_row_style(ws_all_cust.cell(row=c_row, column=8), due_str, align="center")
            apply_row_style(ws_all_cust.cell(row=c_row, column=9), c.get('telegram_id') or "-", align="center")
            c_row += 1
            
    for col in ws_all_cust.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_all_cust.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

    # 3-Varaq: Barcha Amallar va Tranzaksiyalar Tarixi
    ws_tx = wb.create_sheet(title="Amallar Tarixi")
    tx_headers = ["ID", "Sana", "Do'kon Nomi", "Shaxs Ismi", "Amal Turi", "Summa", "Valyuta", "Izoh"]
    for col_idx, h in enumerate(tx_headers, 1):
        apply_header_style(ws_tx.cell(row=1, column=col_idx), h, bg_color="059669")
        
    t_row = 2
    for s in shops:
        txs = await db.get_shop_transactions(s['id'], limit=500)
        for t in txs:
            t_type = "Qarz Qo'shildi" if t['type'] == 'debt' else "To'lov Qilindi"
            curr = t.get('currency', 'UZS') or 'UZS'
            date_str = str(t.get('created_at', ''))[:19]
            
            apply_row_style(ws_tx.cell(row=t_row, column=1), t['id'], align="center")
            apply_row_style(ws_tx.cell(row=t_row, column=2), date_str, align="center")
            apply_row_style(ws_tx.cell(row=t_row, column=3), s['name'])
            apply_row_style(ws_tx.cell(row=t_row, column=4), t.get('customer_name', '-'))
            apply_row_style(ws_tx.cell(row=t_row, column=5), t_type, align="center")
            
            s_cell = ws_tx.cell(row=t_row, column=6)
            s_cell.value = t['amount']
            s_cell.font = Font(name="Arial", size=10)
            s_cell.alignment = Alignment(horizontal="right", vertical="center")
            s_cell.number_format = '#,##0.00' if curr == 'USD' else '#,##0'
            
            apply_row_style(ws_tx.cell(row=t_row, column=7), curr, align="center")
            apply_row_style(ws_tx.cell(row=t_row, column=8), t.get('description') or "-")
            t_row += 1
            
    for col in ws_tx.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_tx.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

async def generate_customer_excel(telegram_id: int, user_full_name: str) -> io.BytesIO:
    """Oddiy mijoz/foydalanuvchi uchun uning shaxsiy barcha qarz va xaridlari Excel hisoboti (So'm va Dollar)"""
    accounts = await db.get_customers_by_telegram_id(telegram_id)
    
    wb = openpyxl.Workbook()
    
    # 1-Varaq: Qayerda qancha qarzim bor
    ws_main = wb.active
    ws_main.title = "Mening Qarzlarim"
    
    headers = ["№", "Do'kon / Shaxs Nomi", "Telefon Raqami", "So'm Qarzi", "Dollar Qarzi ($)"]
    for col_idx, h in enumerate(headers, 1):
        apply_header_style(ws_main.cell(row=1, column=col_idx), h, bg_color="1E3A8A")
        
    total_uzs = 0.0
    total_usd = 0.0
    for row_idx, acc in enumerate(accounts, 2):
        u_val = acc.get('balance', 0.0) or 0.0
        d_val = acc.get('balance_usd', 0.0) or 0.0
        total_uzs += u_val
        total_usd += d_val
        
        apply_row_style(ws_main.cell(row=row_idx, column=1), row_idx - 1, align="center")
        apply_row_style(ws_main.cell(row=row_idx, column=2), acc['shop_name'])
        apply_row_style(ws_main.cell(row=row_idx, column=3), acc['shop_phone'] or "-", align="center")
        apply_row_style(ws_main.cell(row=row_idx, column=4), u_val, align="right", is_money=True)
        
        usd_cell = ws_main.cell(row=row_idx, column=5)
        usd_cell.value = d_val
        usd_cell.font = Font(name="Arial", size=10)
        usd_cell.alignment = Alignment(horizontal="right", vertical="center")
        usd_cell.number_format = '#,##0.00 "$" '
        
    # Jami qatori
    last_row = len(accounts) + 2
    ws_main.cell(row=last_row, column=2, value="JAMI UMUMIY QARZ:").font = Font(name="Arial", size=11, bold=True)
    
    tot_u_cell = ws_main.cell(row=last_row, column=4, value=total_uzs)
    tot_u_cell.font = Font(name="Arial", size=11, bold=True, color="B91C1C")
    tot_u_cell.number_format = '#,##0 "so''m"'
    tot_u_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    tot_d_cell = ws_main.cell(row=last_row, column=5, value=total_usd)
    tot_d_cell.font = Font(name="Arial", size=11, bold=True, color="B91C1C")
    tot_d_cell.number_format = '#,##0.00 "$" '
    tot_d_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    for col in ws_main.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_main.column_dimensions[col[0].column_letter].width = max(max_len + 4, 18)
        
    # 2-Varaq: Xaridlar va To'lovlar Tarixi
    ws_history = wb.create_sheet(title="Xaridlar Tarixi")
    h_headers = ["№", "Sana va Vaqt", "Do'kon / Shaxs Nomi", "Amal Turi", "Summa", "Valyuta", "Tovar / Izoh"]
    for col_idx, h in enumerate(h_headers, 1):
        apply_header_style(ws_history.cell(row=1, column=col_idx), h, bg_color="0D9488")
        
    h_row = 2
    for acc in accounts:
        txs = await db.get_customer_transactions(acc['id'], limit=100)
        for t in txs:
            t_type = "🔴 Qarz olindi" if t['type'] == 'debt' else "🟢 To'lov qilindi"
            curr = t.get('currency', 'UZS') or 'UZS'
            apply_row_style(ws_history.cell(row=h_row, column=1), h_row - 1, align="center")
            apply_row_style(ws_history.cell(row=h_row, column=2), str(t['created_at'])[:19], align="center")
            apply_row_style(ws_history.cell(row=h_row, column=3), acc['shop_name'])
            apply_row_style(ws_history.cell(row=h_row, column=4), t_type, align="center")
            
            s_cell = ws_history.cell(row=h_row, column=5)
            s_cell.value = t['amount']
            s_cell.font = Font(name="Arial", size=10)
            s_cell.alignment = Alignment(horizontal="right", vertical="center")
            s_cell.number_format = '#,##0.00' if curr == 'USD' else '#,##0'
            
            apply_row_style(ws_history.cell(row=h_row, column=6), curr, align="center")
            apply_row_style(ws_history.cell(row=h_row, column=7), t['description'] or "-")
            h_row += 1
            
    for col in ws_history.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_history.column_dimensions[col[0].column_letter].width = max(max_len + 4, 18)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
